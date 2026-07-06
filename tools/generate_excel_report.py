"""
Generate the stakeholder-facing Excel workbook for the E-Commerce Behavior
Analytics project: one sheet per analytics module, same KPIs and headline
findings as the Power BI report, presented as static tables + native Excel
charts. No BigQuery connection required — all figures are hardcoded from the
verified results documented in each sql/0X_*.sql file's header comment block.

Usage:
    python tools/generate_excel_report.py

Output:
    dashboards/ecommerce_analytics.xlsx
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).resolve().parent.parent / "dashboards" / "ecommerce_analytics.xlsx"

# Palette (dataviz skill reference instance) — kept consistent with the notebooks
NAVY = "1A2E4A"
BLUE = "2A78D6"
LIGHT_BLUE = "CDE2FB"
WHITE = "FFFFFF"
GRAY_TEXT = "52514E"
GRID = "E1E0D9"

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=WHITE)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color=GRAY_TEXT)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
BODY_FONT = Font(name="Calibri", size=10, color="0B0B0B")
KPI_LABEL_FONT = Font(name="Calibri", size=10, color=GRAY_TEXT)
KPI_VALUE_FONT = Font(name="Calibri", size=14, bold=True, color=NAVY)

TITLE_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FILL = PatternFill("solid", fgColor=BLUE)
STRIPE_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
KPI_FILL = PatternFill("solid", fgColor="F9F9F7")

THIN_GRAY = Side(style="thin", color=GRID)
CELL_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def write_title(ws, title, subtitle):
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = title
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    sub = ws["A2"]
    sub.value = subtitle
    sub.font = SUBTITLE_FONT
    sub.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[2].height = 18


def write_kpi_row(ws, row, kpis):
    """kpis: list of (label, value) tuples, up to 4 per row, each spanning 2 columns."""
    col = 1
    for label, value in kpis:
        start = get_column_letter(col)
        end = get_column_letter(col + 1)
        ws.merge_cells(f"{start}{row}:{end}{row}")
        lbl_cell = ws[f"{start}{row}"]
        lbl_cell.value = label
        lbl_cell.font = KPI_LABEL_FONT
        lbl_cell.fill = KPI_FILL
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"{start}{row + 1}:{end}{row + 1}")
        val_cell = ws[f"{start}{row + 1}"]
        val_cell.value = value
        val_cell.font = KPI_VALUE_FONT
        val_cell.fill = KPI_FILL
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        col += 2
    ws.row_dimensions[row].height = 16
    ws.row_dimensions[row + 1].height = 24
    return row + 3


def write_table(ws, start_row, headers, rows, number_formats=None):
    """Writes a header row + data rows starting at start_row. Returns the row after the table."""
    number_formats = number_formats or {}
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER

    for r, row_data in enumerate(rows, start=start_row + 1):
        striped = (r - start_row) % 2 == 0
        for c, value in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = BODY_FONT
            cell.border = CELL_BORDER
            cell.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")
            if striped:
                cell.fill = STRIPE_FILL
            if c in number_formats:
                cell.number_format = number_formats[c]

    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = max(14, len(str(headers[c - 1])) + 4)

    return start_row + len(rows) + 2


def add_bar_chart(ws, anchor_cell, data_min_col, data_max_col, cat_col, min_row, max_row,
                   title, y_title, height=8, width=16):
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.y_axis.title = y_title
    chart.style = 10
    chart.height = height
    chart.width = width

    data = Reference(ws, min_col=data_min_col, max_col=data_max_col, min_row=min_row - 1, max_row=max_row)
    cats = Reference(ws, min_col=cat_col, max_col=cat_col, min_row=min_row, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor_cell)


def add_line_chart(ws, anchor_cell, data_min_col, data_max_col, cat_col, min_row, max_row,
                    title, y_title, height=8, width=16):
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_title
    chart.style = 10
    chart.height = height
    chart.width = width

    data = Reference(ws, min_col=data_min_col, max_col=data_max_col, min_row=min_row - 1, max_row=max_row)
    cats = Reference(ws, min_col=cat_col, max_col=cat_col, min_row=min_row, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    for s in chart.series:
        s.smooth = False
        s.marker.symbol = "circle"
    ws.add_chart(chart, anchor_cell)


# ============================================================
# Sheet 0: Overview
# ============================================================
def build_overview(wb):
    ws = wb.active
    ws.title = "Overview"
    write_title(ws, "E-Commerce Behavior Analytics", "REES46 Open CDP  |  411,709,736 events  |  Oct 2019 - Apr 2020  |  Kazakhstan / CIS market")

    row = write_kpi_row(ws, 4, [
        ("Total Events", "411,709,736"),
        ("Purchasing Users", "2,064,899"),
        ("Overall Conversion", "6.09%"),
        ("Total Revenue (all categories)", "$2.06B"),
    ])
    row = write_kpi_row(ws, row, [
        ("Cart Abandonment Rate", "54.09%"),
        ("Champion Segment Revenue Share", "41.40%"),
        ("Month-1 Retention (avg)", "19.36%"),
        ("COVID Session Conversion Lift", "+18.0%"),
    ])

    row += 1
    ws.cell(row=row, column=1, value="Module Index").font = Font(bold=True, size=12, color=NAVY)
    row += 1
    headers = ["Module", "Business Question", "Key Finding"]
    modules = [
        ("1. Funnel Analysis", "What is view->cart->purchase conversion?", "6.09% overall; 'construction' category leads at 8.24% (see Module 5 taxonomy note)"),
        ("2. Session Analytics", "How do users browse before buying?", "Conversion plateaus after 6 events/session; median purchase happens at event #5"),
        ("3. RFM Segmentation", "Who are the best customers?", "Champions: 13.2% of users generate 41.4% of revenue"),
        ("4. Cohort Retention", "Do customers come back?", "Month-1 retention averages 19.4% - the single biggest drop in the customer lifecycle"),
        ("5. Category & Brand", "What drives revenue?", "Apple is the top brand by revenue ($929M); 'construction' taxonomy shift Dec 2019"),
        ("6. Anomaly Detection", "What looks unusual?", "Platform price cap at $2,574.07; confirmed logging gap Feb 27, 2020"),
        ("7. COVID Quasi-Experiment", "Did COVID change behavior?", "Session conversion +18% (p<<0.001); AOV fell 11.7%; retention catalyst not acquisition driver"),
    ]
    write_table(ws, row, headers, modules)
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False


# ============================================================
# Sheet 1: Module 1 - Funnel Analysis
# ============================================================
def build_funnel(wb):
    ws = wb.create_sheet("01 Funnel")
    write_title(ws, "Module 1: Funnel Analysis", "What is the conversion rate from view -> cart -> purchase, and where does it break down?")

    row = write_kpi_row(ws, 4, [
        ("Sessions with View", "89,479,619"),
        ("Overall Conversion Rate", "6.09%"),
        ("Cart Abandonment Rate", "54.09%"),
        ("Direct-Purchase Sessions", "530,292"),
    ])

    row += 1
    ws.cell(row=row, column=1, value="Funnel by Category (top 10 by conversion)").font = Font(bold=True, color=NAVY)
    row += 1
    cat_headers = ["Category (raw)", "Views", "Carts", "Purchases", "Conversion %"]
    cat_rows = [
        ["construction", 23052672, 3462460, 1898778, 0.0824],
        ["electronics", 19472466, 2061398, 1130860, 0.0581],
        ["appliances", 15715297, 1615685, 786186, 0.0500],
        ["sport", 5712881, 593960, 265666, 0.0465],
        ["unknown", 16804172, 1395989, 676904, 0.0403],
        ["auto", 1457937, 93210, 48078, 0.0330],
        ["furniture", 5466667, 403192, 172881, 0.0316],
        ["computers", 5700141, 362813, 173853, 0.0305],
        ["apparel", 12854885, 929354, 388777, 0.0302],
        ["country_yard", 345259, 26255, 9532, 0.0276],
    ]
    cat_start = row
    row = write_table(ws, row, cat_headers, cat_rows,
                       number_formats={2: "#,##0", 3: "#,##0", 4: "#,##0", 5: "0.00%"})
    add_bar_chart(ws, f"G{cat_start}", 5, 5, 1, cat_start + 1, cat_start + len(cat_rows),
                  "Conversion Rate by Category", "Conversion %")

    row += 1
    ws.cell(row=row, column=1, value="Funnel by Brand (top 8 by purchase volume)").font = Font(bold=True, color=NAVY)
    row += 1
    brand_headers = ["Brand", "Views", "Purchases", "Overall Conv %", "Cart->Purchase %"]
    brand_rows = [
        ["samsung", 16383360, 1279265, 0.0781, 0.5742],
        ["apple", 12545341, 997688, 0.0795, 0.5700],
        ["xiaomi", 8975402, 457372, 0.0510, 0.4752],
        ["huawei", 3534417, 190515, 0.0539, 0.5596],
        ["oppo", 1670660, 99562, 0.0596, 0.5810],
        ["lucente", 2017765, 91893, 0.0455, 0.6396],
        ["lg", 1910874, 77650, 0.0406, 0.4886],
        ["sony", 1915499, 65243, 0.0341, 0.4674],
    ]
    write_table(ws, row, brand_headers, brand_rows,
                number_formats={2: "#,##0", 3: "#,##0", 4: "0.00%", 5: "0.00%"})

    ws.sheet_view.showGridLines = False


# ============================================================
# Sheet 2: Module 2 - Session Analytics
# ============================================================
def build_sessions(wb):
    ws = wb.create_sheet("02 Sessions")
    write_title(ws, "Module 2: Session Analytics", "How do users browse before they buy?")

    row = write_kpi_row(ws, 4, [
        ("Total Sessions", "89,693,595"),
        ("Median Events / Session", "2"),
        ("Median Session Duration", "42 sec"),
        ("Median Event # at Purchase", "5"),
    ])

    row += 1
    ws.cell(row=row, column=1, value="Conversion by Session Depth Bucket").font = Font(bold=True, color=NAVY)
    row += 1
    headers = ["Event Bucket", "Total Sessions", "Purchasing Sessions", "Conversion %"]
    rows_ = [
        ["1 event", 36377146, 53442, 0.0015],
        ["2-5 events", 32911072, 2518249, 0.0765],
        ["6-10 events", 11260308, 1596659, 0.1418],
        ["11-20 events", 6259229, 872343, 0.1394],
        ["21-50 events", 2559907, 359341, 0.1404],
        ["51+ events", 325933, 49900, 0.1531],
    ]
    start = row
    row = write_table(ws, row, headers, rows_,
                       number_formats={2: "#,##0", 3: "#,##0", 4: "0.00%"})
    add_bar_chart(ws, f"F{start}", 4, 4, 1, start + 1, start + len(rows_),
                  "Conversion Rate by Session Depth", "Conversion %")

    row += 1
    ws.cell(row=row, column=1, value="Converting vs. Non-Converting Sessions").font = Font(bold=True, color=NAVY)
    row += 1
    headers2 = ["Session Type", "Sessions", "Avg Events", "Median Events", "Avg Carts", "Median Duration (min)"]
    rows2 = [
        ["Converting", 5449934, 8.90, 6, 1.74, 4.38],
        ["Non-Converting", 84243661, 4.31, 2, 0.11, 0.50],
    ]
    write_table(ws, row, headers2, rows2,
                number_formats={2: "#,##0", 3: "0.00", 5: "0.00", 6: "0.00"})

    ws.sheet_view.showGridLines = False


# ============================================================
# Sheet 3: Module 3 - RFM Segmentation
# ============================================================
def build_rfm(wb):
    ws = wb.create_sheet("03 RFM")
    write_title(ws, "Module 3: RFM Segmentation", "Who are the best customers, and how concentrated is revenue among them?")

    row = write_kpi_row(ws, 4, [
        ("Purchasing Users", "2,064,899"),
        ("Champion Users", "273,211 (13.23%)"),
        ("Champion Revenue Share", "41.40%"),
        ("At Risk Users (win-back target)", "419,470"),
    ])

    row += 1
    ws.cell(row=row, column=1, value="Segment Distribution").font = Font(bold=True, color=NAVY)
    row += 1
    headers = ["Segment", "% Users", "% Revenue", "Avg Recency (days)", "Avg Frequency", "Avg Monetary"]
    rows_ = [
        ["Champion", 0.1323, 0.4140, 27.8, 9.24, 3117.85],
        ["Loyal", 0.2645, 0.2566, 54.7, 3.69, 966.62],
        ["At Risk", 0.2031, 0.2218, 152.6, 3.53, 1087.71],
        ["Others", 0.2064, 0.0572, 130.2, 1.00, 276.00],
        ["Recent but Infrequent", 0.1400, 0.0356, 27.3, 1.00, 253.60],
        ["Lost", 0.0536, 0.0148, 182.4, 1.00, 274.59],
    ]
    start = row
    row = write_table(ws, row, headers, rows_,
                       number_formats={2: "0.00%", 3: "0.00%", 4: "0.0", 5: "0.00", 6: "$#,##0.00"})
    add_bar_chart(ws, f"H{start}", 2, 3, 1, start + 1, start + len(rows_),
                  "Users % vs. Revenue % by Segment", "Share of total")

    row += 1
    ws.cell(row=row, column=1, value="Monetary Distribution by Segment").font = Font(bold=True, color=NAVY)
    row += 1
    headers2 = ["Segment", "P25", "Median", "P75", "P90", "Max"]
    rows2 = [
        ["Champion", 792.74, 1384.98, 2869.93, 6210.49, 790120.94],
        ["At Risk", 211.68, 465.34, 1081.29, 2366.60, 473119.39],
        ["Loyal", 180.07, 361.06, 830.60, 2002.10, 474648.46],
        ["Others", 82.33, 170.15, 311.08, 733.30, 2574.07],
        ["Lost", 77.22, 169.35, 308.86, 719.68, 2574.07],
        ["Recent but Infrequent", 64.33, 166.77, 302.08, 591.78, 2574.07],
    ]
    write_table(ws, row, headers2, rows2,
                number_formats={2: "$#,##0.00", 3: "$#,##0.00", 4: "$#,##0.00", 5: "$#,##0.00", 6: "$#,##0.00"})

    ws.sheet_view.showGridLines = False


# ============================================================
# Sheet 4: Module 4 - Cohort Retention
# ============================================================
def build_cohort(wb):
    ws = wb.create_sheet("04 Cohort")
    write_title(ws, "Module 4: Cohort Retention", "Do customers come back month after month?")

    row = write_kpi_row(ws, 4, [
        ("Cohorts Tracked", "7 (Oct 2019 - Apr 2020)"),
        ("Avg M1 Retention", "19.36%"),
        ("Avg M6 Retention", "8.08%"),
        ("Largest Cohort", "Apr 2020 (320,535)"),
    ])

    row += 1
    ws.cell(row=row, column=1, value="Cohort Retention Matrix (blank = outside observation window, not zero)").font = Font(bold=True, color=NAVY)
    row += 1
    headers = ["Cohort", "Cohort Users", "M0", "M1", "M2", "M3", "M4", "M5", "M6"]
    rows_ = [
        ["Oct 2019", 347118, 1.00, 0.263, 0.219, 0.137, 0.127, 0.122, 0.081],
        ["Nov 2019", 350352, 1.00, 0.221, 0.122, 0.115, 0.111, 0.077, None],
        ["Dec 2019", 347286, 1.00, 0.152, 0.125, 0.120, 0.081, None, None],
        ["Jan 2020", 215886, 1.00, 0.184, 0.136, 0.088, None, None, None],
        ["Feb 2020", 225048, 1.00, 0.190, 0.104, None, None, None, None],
        ["Mar 2020", 258674, 1.00, 0.151, None, None, None, None, None],
        ["Apr 2020", 320535, 1.00, None, None, None, None, None, None],
    ]
    write_table(ws, row, headers, rows_,
                number_formats={2: "#,##0", 3: "0.0%", 4: "0.0%", 5: "0.0%", 6: "0.0%", 7: "0.0%", 8: "0.0%", 9: "0.0%"})

    row += len(rows_) + 3
    ws.cell(row=row, column=1, value="Average Retention Curve").font = Font(bold=True, color=NAVY)
    row += 1
    headers2 = ["Months Since First Purchase", "Avg Retention", "Min", "Max"]
    rows2 = [
        ["M0", 1.00, 1.00, 1.00],
        ["M1", 0.1936, 0.1513, 0.2630],
        ["M2", 0.1412, 0.1038, 0.2193],
        ["M3", 0.1149, 0.0879, 0.1373],
        ["M4", 0.1063, 0.0812, 0.1270],
        ["M5", 0.0994, 0.0767, 0.1221],
        ["M6", 0.0808, 0.0808, 0.0808],
    ]
    start2 = row
    row = write_table(ws, row, headers2, rows2,
                       number_formats={2: "0.0%", 3: "0.0%", 4: "0.0%"})
    add_line_chart(ws, f"F{start2}", 2, 2, 1, start2 + 1, start2 + len(rows2),
                   "Average Retention Curve", "Retention %")

    ws.sheet_view.showGridLines = False


# ============================================================
# Sheet 5: Module 5 - Category & Brand Performance
# ============================================================
def build_category_brand(wb):
    ws = wb.create_sheet("05 Category & Brand")
    write_title(ws, "Module 5: Category & Brand Performance", "What drives revenue and where is demand concentrated?")

    row = write_kpi_row(ws, 4, [
        ("Total Platform Revenue", "~$2.06B"),
        ("Top Brand by Revenue", "Apple ($929M)"),
        ("Top Brand by Volume", "Samsung (1.57M purchases)"),
        ("Most Premium Brand", "Thermomix (avg $1,683)"),
    ])

    row += 1
    ws.cell(row=row, column=1, value="Category Performance (display names; 'DIY/Home Improvement' = raw 'construction' - see taxonomy note)").font = Font(bold=True, color=NAVY)
    row += 1
    headers = ["Category", "Views", "Purchases", "Total Revenue", "Avg Price"]
    rows_ = [
        ["DIY / Home Improvement *", 77482657, 2442248, 1008523959, 412.95],
        ["Electronics", 69201207, 1341139, 511171792, 381.15],
        ["Appliances", 63417883, 925763, 237020688, 256.03],
        ["Unknown / No Category", 62105978, 768060, 99842875, 129.99],
        ["Apparel", 41137209, 455370, 62552043, 137.37],
        ["Computers & Peripherals", 18733089, 202396, 49987328, 246.98],
        ["Sports & Outdoor", 16145807, 316243, 41518787, 131.29],
        ["Furniture & Home", 17652355, 197485, 23122285, 117.08],
    ]
    start = row
    row = write_table(ws, row, headers, rows_,
                       number_formats={2: "#,##0", 3: "#,##0", 4: "$#,##0", 5: "$#,##0.00"})
    add_bar_chart(ws, f"G{start}", 4, 4, 1, start + 1, start + len(rows_),
                  "Revenue by Category", "Total Revenue ($)")

    row += 1
    ws.cell(row=row, column=1, value="* Taxonomy note: 'construction' contains Apple/Samsung/Xiaomi smartphones,").font = SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value="  misclassified starting Dec 2, 2019. See README Data Quality section.").font = SUBTITLE_FONT
    row += 2

    ws.cell(row=row, column=1, value="Top 10 Brands by Revenue").font = Font(bold=True, color=NAVY)
    row += 1
    brand_headers = ["Brand", "Purchases", "Total Revenue", "Avg Price", "Categories"]
    brand_rows = [
        ["apple", 1246326, 929321587, 745.65, 11],
        ["samsung", 1567074, 425423961, 271.48, 12],
        ["xiaomi", 542848, 91406924, 168.38, 13],
        ["huawei", 227722, 42203598, 185.33, 10],
        ["lg", 91108, 38260344, 419.94, 9],
        ["acer", 61939, 31996239, 516.58, 8],
        ["sony", 75995, 28528875, 375.40, 12],
        ["lucente", 108910, 28410590, 260.86, 10],
        ["oppo", 119433, 26428298, 221.28, 3],
        ["lenovo", 57028, 22197927, 389.25, 9],
    ]
    write_table(ws, row, brand_headers, brand_rows,
                number_formats={2: "#,##0", 3: "$#,##0", 4: "$#,##0.00"})

    ws.sheet_view.showGridLines = False


# ============================================================
# Sheet 6: Module 6 - Anomaly Detection
# ============================================================
def build_anomaly(wb):
    ws = wb.create_sheet("06 Anomaly Detection")
    write_title(ws, "Module 6: Anomaly Detection", "What purchase patterns and data signals look unusual?")

    row = write_kpi_row(ws, 4, [
        ("Platform Price Cap", "$2,574.07"),
        ("Confirmed Logging Gap", "Feb 27, 2020"),
        ("Bot Sessions Detected", "374 (>500 events)"),
        ("Highest Traffic Day", "Nov 15, 2019 (6.2M events)"),
    ])

    row += 1
    ws.cell(row=row, column=1, value="Price Outlier Summary (IQR method) - top categories by outlier %").font = Font(bold=True, color=NAVY)
    row += 1
    headers = ["Category (raw)", "Total Purchases", "Upper Fence", "Max Price", "Outlier Count", "Outlier %"]
    rows_ = [
        ["stationery", 1872, 167.12, 942.88, 269, 0.1437],
        ["apparel", 455370, 336.67, 2557.59, 58771, 0.1291],
        ["furniture", 197485, 246.22, 2574.07, 20521, 0.1039],
        ["unknown", 768060, 309.52, 2574.07, 79648, 0.1037],
        ["country_yard", 10849, 78.86, 859.74, 977, 0.0901],
        ["kids", 93757, 277.72, 2574.04, 7525, 0.0803],
        ["computers", 202396, 619.38, 2574.04, 18572, 0.0918],
        ["electronics", 1341139, 1039.76, 2574.07, 78618, 0.0586],
        ["construction", 2442248, 1298.49, 2574.07, 118608, 0.0486],
        ["appliances", 925763, 803.73, 2574.04, 27724, 0.0300],
    ]
    write_table(ws, row, headers, rows_,
                number_formats={2: "#,##0", 3: "$#,##0.00", 4: "$#,##0.00", 5: "#,##0", 6: "0.00%"})

    row += len(rows_) + 3
    ws.cell(row=row, column=1, value="Key Findings").font = Font(bold=True, color=NAVY)
    row += 1
    findings = [
        ["Price cap confirmed", "All major categories hit an identical max price ($2,574.07) - a platform-level price ceiling, not fraud."],
        ["Bot detection", "374 sessions with >500 events; top 10 all show 0 purchases, views only - confirmed scrapers."],
        ["Logging gap", "Feb 27, 2020: 197,047 events vs ~2M normal (z=-18.84) - the only confirmed data outage."],
        ["Taxonomy shift", "'construction' category revenue jumped from <$1.1M/month to $217M in the week of Dec 2, 2019 - Apple/Samsung/Xiaomi reclassified from electronics."],
    ]
    write_table(ws, row, ["Finding", "Detail"], findings)
    ws.column_dimensions["B"].width = 90

    ws.sheet_view.showGridLines = False


# ============================================================
# Sheet 7: Module 7 - COVID Quasi-Experiment
# ============================================================
def build_covid(wb):
    ws = wb.create_sheet("07 COVID")
    write_title(ws, "Module 7: COVID Quasi-Experiment", "Did the Kazakhstan COVID-19 lockdown (Mar 16, 2020) measurably change purchasing behaviour?")

    row = write_kpi_row(ws, 4, [
        ("Session Conversion Lift", "+18.0% (p<<0.001)"),
        ("AOV Change", "-11.7%"),
        ("Returning Buyer Daily Rate", "+71%"),
        ("New Buyer Daily Rate", "-6% (flat)"),
    ])

    row += 1
    ws.cell(row=row, column=1, value="Pre-COVID vs. COVID Onset").font = Font(bold=True, color=NAVY)
    row += 1
    headers = ["Metric", "Pre-COVID", "COVID Onset", "Change"]
    rows_ = [
        ["Session Conversion %", 0.0584, 0.0689, "+18.0%"],
        ["Avg Order Value ($)", 307.49, 271.67, "-11.7%"],
        ["Avg Events / Session", 4.39, 5.50, "+25.3%"],
        ["Avg Products Viewed", 2.79, 3.41, "+22.2%"],
        ["Avg Session Revenue ($)", 389.23, 333.68, "-14.3%"],
    ]
    start = row
    row = write_table(ws, row, headers, rows_,
                       number_formats={2: "General", 3: "General"})
    add_bar_chart(ws, f"F{start}", 2, 3, 1, start + 1, start + 1,
                  "Session Conversion: Pre vs. COVID", "Conversion", height=6, width=12)

    row += 1
    ws.cell(row=row, column=1, value="Category Revenue Mix Shift (Jan 2020 baseline; DIY/construction excluded - taxonomy artifact)").font = Font(bold=True, color=NAVY)
    row += 1
    cat_headers = ["Category", "Pre-COVID Share %", "COVID Share %", "Delta (pp)"]
    cat_rows = [
        ["Appliances", 0.1106, 0.1645, 5.4],
        ["Electronics", 0.0719, 0.1172, 4.5],
        ["Apparel", 0.0344, 0.0469, 1.2],
        ["Unknown / No Category", 0.0238, 0.0443, 2.0],
        ["Sports & Outdoor", 0.0284, 0.0201, -0.8],
        ["Computers & Peripherals", 0.0154, 0.0111, -0.4],
    ]
    write_table(ws, row, cat_headers, cat_rows,
                number_formats={2: "0.0%", 3: "0.0%"})

    ws.sheet_view.showGridLines = False


def main():
    wb = Workbook()
    build_overview(wb)
    build_funnel(wb)
    build_sessions(wb)
    build_rfm(wb)
    build_cohort(wb)
    build_category_brand(wb)
    build_anomaly(wb)
    build_covid(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Workbook saved to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
